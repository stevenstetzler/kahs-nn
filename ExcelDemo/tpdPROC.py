#!/usr/bin/env python
"""!/usr/bin/env openpyxl

K that seemed to fix it.

Other thing:

The background line fit seems a bit underestimated for most of them.

Maybe one thing to add:

At the end, you can plot all of the Ch 1 - background data points on the same plot.
Make sure the data is all structured so if I change the background fit it will adjust these curves too.


"""

#Josh wants you to make data points visible, add red peak
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
import os
import os.path
import math

#constants: Row and column values of all relevant labels
maxrow = 150
minrow = 5
UnitRow = 3
TempCol = 3
Ch1Col = 4
Ch2Col = 5
Ch3Col = 6
BckgrdCol = 7
BckCh1Col = 8
rnge = 8
before_p = 0
after_p = 0
before_t = 0
after_t = 0
high_ind = 0
buffr = .30

class Datap:
    def __init__(self):
        self.pres = 0
        self.temp = 0


#Declare spreadsheet, and input location of folder
path = raw_input('TPD Background Analyzer \nFilepath of spreadsheet folder: ')
exp_peak = raw_input('What is the expected peak temperature?')
wb = Workbook()

#Copies element of one sheet to the same space in another sheet
def copy_elems(row, col, ws, tempsheet):
    let = chr(col+64)
    num = str(row)
    ws.cell( column=col, row=row, value=tempsheet[let + str(row)].value)

#Iterates through an implementer defined range of a spreadsheet, and copies the
# elements into a new sheet by calling copy_elems
def copy_sheet(wb, tempsheet, file):
    #removes xlsx tag
    if file.endswith('.xlsx'):
        file = file[:-5]
    ws = wb.create_sheet(file)
    for row in range(1, maxrow):
        for col in range(1,Ch3Col+1):
            copy_elems(row, col, ws, tempsheet)
    return ws

#Generates a list of values from those stored in Channel 1 of the excel sheet
def col_A_vals(ws):
    values = []
    num = 0
    for row in range(minrow, maxrow):
        if str(ws['D'+ str(row)].value) != "None":
            temp = Datap()
            temp.temp = float(ws[chr(Ch1Col+63)+ str(row)].value)
            temp.pres = float(ws[chr(Ch1Col+64)+ str(row)].value)
            values.append(temp)
    global high_ind
    high_ind = (len(values)+4)
    return values

#Fills background calculations
def fill_background(ws, i):
    ws.cell(column=BckgrdCol, row= minrow + i, value= "=K$4*C" + str(minrow+i) + "+L$4")
    ws.cell(column=BckCh1Col, row= minrow + i, value= "=D" + str(5+i) + "-G" + str(minrow+i))

#Fils linear fit/integration calculations


def fill_formulas(i, start, ws):
    ws.cell(column=9, row= start+i, value= "=C" + str(start+i) + "-C" + str(start+i-1))
    ws.cell(column=10, row= start+i, value= "=(H" + str(start+i) + "+H" + str(start+i-1) + ")/2")
    ws.cell(column=11, row= start+i, value= "=I" + str(start+i) + "*J" + str(start+i))
    ws.cell(column=12, row= start+i, value= "=AVERAGE(C" + str(start+i-1) + ":C" + str(start+i) + ")")

#Uses an algorithm to find the maxima of the TPD curve. This algorithm only
# detects maxima in the first 65% of the range, and stops searching when the
# value succeeding the next value is less than the next value. I can see how
# this could go wrong, but as long as you make sure you tak data for a signi-
# ficant time after you find the peak, you should be good to go
#Finds value, then finds index of that value in the values list
def find_max_index(array):
    val = int(exp_peak)
    ind_guess = 0
    ind = 0
    for i in range(len(array)-1):
        if val >= array[i].temp and val <= array[i+1].temp:
            break
        ind_guess += 1
    ind = ind_guess
    val = array[ind].pres
    for i in range(int(len(array)*.165)):
        if ind_guess + i < len(array):
            if val < array[ind_guess + i].pres:
                ind = ind_guess + i
                val = array[ind].pres
        if ind_guess - i > 0:
            if val < array[ind_guess - i].pres:
                ind = ind_guess - i
                val = array[ind].pres
    print( "Peak found at " + str(val))
    return ind

def divide(dx, dy):
    if dx == 0:
        return dy
    else:
        return dy/dx

def comp_derivs(array, peak, i, k):
    dy = (-1*k*(array[peak+k*i].pres) + k*array[peak+k*(i+1)].pres)
    dx = (-1*k*(array[peak+k*i].temp) + k*array[peak+k*(i+1)].temp)
    m_curr = divide(dx, dy)
    dy = (-1*k*(array[peak+k*(i+1)].pres) + k*array[peak+k*(i+2)].pres)
    dx = (-1*k*(array[peak+k*(i+1)].temp) + k*array[peak+k*(i+2)].temp)
    m_next = divide(dx, dy)
    return [m_curr, m_next]

def find_bounds(array, peak):
    bounds = [0 for i in range(2)]
    left = False; right = False
    lbuf = int(len(array)*buffr); rbuf = lbuf
    i = int(.15*len(array))
    while not(left and right):
        if not left:
            if peak-i <= 0:
                bounds[0] = 0
                left = True
            elif right:
                lbuf = lbuf-1
                if lbuf <= 0 and array[peak-i].pres < array[bounds[1]].pres:
                    left = True
                    bounds[0] = peak-i
                    continue
            if peak-i >= 2:
                derivs = comp_derivs(array, peak, i, -1)
                if derivs[0]*derivs[1] < 0:
                    left = True
                    bounds[0] = peak-i+1
        if not right:
            if peak+i >= len(array)-1:
                bounds[1] = len(array)-1
                right = True
            elif left:
                rbuf = rbuf-1
                if rbuf <= 0 and array[peak+i].pres < array[bounds[0]].pres:
                    right = True
                    bounds[1] = peak+i
                    continue
            if peak+i < len(array)-2:
                derivs = comp_derivs(array, peak, i, 1)
                if  derivs[0]/derivs[1] < 0:
                    right = True
                    bounds[1] = peak+i

        i += 1
        print(left)
        print(right)
    return bounds

#Handles all calculations relevant for finding background fit. First calculates
# background line, then performs all background calculations
def background(ws):
    array = col_A_vals(ws)
    max_index = find_max_index(array)
    #find_peak_bounds(max_index, array)
    #Does the before and after calculations for he linear fit
    bounds = find_bounds(array, max_index)

    print (bounds)
    print (len(array))
    print (max_index)


    before_t = array[bounds[0]].temp
    after_t = array[bounds[1]].temp
    before_p = array[bounds[0]].pres
    after_p = array[bounds[1]].pres

    #Fills in background and ch-1 background columns into the table
    ws.cell(column=BckgrdCol, row=UnitRow, value="Background")
    ws.cell(column=BckCh1Col, row=UnitRow, value="Ch-1 Background")
    ws.cell(column=10, row= 2, value= "before")
    ws.cell(column=10, row= 3, value= "after")
    ws.cell(column=10, row= 4, value= "Curve Area")
    ws.cell(column=11, row= 1, value= "T")
    ws.cell(column=12, row= 1, value= "P")
    ws.cell(column=11, row= 2, value= before_t)
    ws.cell(column=12, row= 2, value= before_p)
    ws.cell(column=11, row= 3, value= after_t)
    ws.cell(column=12, row= 3, value= after_p)
    ws.cell(column=10, row= 4, value= "Line Fit, m, b")
    ws.cell(column=11, row= 4, value= "=INDEX(LINEST(L2:L3,K2:K3,1),1)")
    ws.cell(column=12, row= 4, value= "=INDEX(LINEST(L2:L3,K2:K3,1),2)")
    print("Background line computed")
    #Fills in formulae
    for i in range(len(array)):
        fill_background(ws, i)
    print( "Background information calculated")

    #Fills in the labels for the integration calculation. Calls fill_formulas to do calculations
    start = UnitRow+2+bounds[0]
    ws.cell(column=9, row= start-1, value="x2-x1")
    ws.cell(column=10, row= start-1, value="(y2+y1)/2")
    ws.cell(column=11, row= start-1, value="Area")
    ws.cell(column=12, row= start-1, value="Avg x")
    ws.cell(column=10, row=6, value="Total Area")


    integ_range = bounds[1]-bounds[0]+1
    for i in range(integ_range):
        fill_formulas(i, start, ws)
    ws.cell(column=11, row=6, value="=(SUMIF(K" + str(start) + ":K" + str(integ_range) + ', ">0"))')
    print( "Integration formulas calculated")

    return [start, bounds[1] + bounds[0]+1]

#Creates the raw data and adjusted data scatter plots
def create_simple_charts(ws, start, boundl):
    #The Raw data TPD chart - Declares x values and the chart variables
    xvalues = Reference(ws, min_col=TempCol, min_row=minrow, max_row=maxrow)
    chart = ScatterChart(scatterStyle='lineMarker')
    #Fills Y values (using the reference function). Appends the series to the chart.
    #Repeats for all three curvess
    for i in range(Ch1Col,Ch3Col+1):
        values = Reference(ws, min_col=i, min_row=minrow, max_row=maxrow)
        series = Series(values, xvalues, title_from_data=True)
        chart.series.append(series)
        s1 = chart.series[i-4]
        s1.marker.symbol = "diamond"

    xvalues = Reference(ws, min_col=11, min_row=2, max_row=3)
    values = Reference(ws, min_col=12, min_row=2, max_row=3)
    series= Series(values, xvalues)
    chart.series.append(series)

    #Labelling
    chart.title = "TPD"
    chart.x_axis.title = "Temperature (K)"
    chart.y_axis.title = "Pressure"
    ws.add_chart(chart, "M6")

    #Adjuested data TPD values - identical process as above
    chart = ScatterChart(scatterStyle='lineMarker')
    xvalues = Reference(ws, min_col=TempCol, min_row=minrow, max_row=high_ind)
    values = Reference(ws, min_col=BckCh1Col, min_row=minrow, max_row=high_ind)
    series = Series(values, xvalues, title_from_data=True)
    chart.series.append(series)
    s1 = chart.series[0]
    s1.marker.symbol = "diamond"

    #Integrated TPD values
    xvalues = Reference(ws, min_col=12, min_row=start, max_row=start+boundl)
    values = Reference(ws, min_col=10, min_row=start, max_row=start+boundl)
    series = Series(values, xvalues, title_from_data=True)
    chart.series.append(series)
    s1 = chart.series[1]
    s1.marker.symbol = "diamond"

    chart.title = "TPD"
    chart.x_axis.title = "Temperature (K)"
    chart.y_axis.title = "Pressure"
    ws.add_chart(chart, "M37")
    return wb

#Main
def main():
    files_found = 0
    #Check for valid path
    if os.path.exists(path):
        print("Path valid... Searching for valid files")
        for file in os.listdir(path):
            if file.startswith("~$"):
                continue
            if file.endswith(".xlsx"):
                #Loads workbook to be copied, retrieves the sheet, and copies
                #sheet to the new workbook
                tempb = load_workbook(path + "/" + file)
                tempsheet = tempb['Sheet1']
                ws = copy_sheet(wb, tempsheet, file)
                print("\n" + file + " Successfully copied" + "\n")
                #Fill in background information
                start = background(ws)
                print("\n Background analysis completed... Inserting charts")
                #Create relevant charts
                create_simple_charts(ws, start[0], start[1])
                print("Charts inserted \n")
                files_found += 1
        if files_found == 1:
            print("No excel files found")
        else:
            print("\nFound and processed " + str(files_found) + " files")
    else:
        print("Invalid directory")

main()
name = raw_input('Output filename: ')
wb.save(name + ".xlsx")
